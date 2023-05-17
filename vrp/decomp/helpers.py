# Author: Xu Ye <kan.ye@tum.de>

import os
import argparse
import math
import time
from typing import Callable
import json
from functools import lru_cache

import numpy as np
import pandas as pd
import openpyxl as xl
from ..third_party import cvrplib

from .decomposition import Node, VRPInstance
from .logger import logger
from .constants import *


def get_args_parser(script_name):
    # Usage example: args = helpers.get_args_parser(os.path.basename(__file__))
    parser = argparse.ArgumentParser(
        description="Example usages: "
            f"python {script_name} -b=1 -n='C206' -k=3 -t | "
            f"python {script_name} -b=2 -n='C1_2_1' -k=3 -t"
    )

    parser.add_argument(
        '-n', '--instance_name', required=True,
        help='benchmark instance name without file extension, e.g. "C206"'
    )
    parser.add_argument(
        '-b', '--benchmark', default=1, choices=[1, 2], type=int,
        help='benchmark dataset to use: 1=Solomon (1987), '
            '2=Homberger and Gehring (1999); Default=1'
    )
    parser.add_argument('-k', '--num_clusters', required=True, type=int,
                        help='number of clusters')
    parser.add_argument('-t', '--include_time_windows', action='store_true',
                        help='use time windows for clustering or not')

    args = parser.parse_args()
    return args


def get_min_tours(demands, capacity):
    """Get the minimum number of tours (i.e. vehicles required)."""
    # total demand of all customers / vehicle capacity
    return math.ceil(sum(demands) / capacity)


def standardize_feature_vectors(fv):
    """Standardize feature vectors using z-score."""
    if not isinstance(fv, np.ndarray):
        fv = np.array(fv)
    # axis=0 -> row axis, runs down the rows, i.e. calculate the mean for each column/feature
    # aggregate over rows, within each column
    mean = np.mean(fv, axis=0)
    # ddof=1 -> degrees of freedom = N-1, i.e. sample std
    # ddof = 'delta degrees of freedom'
    # set ddof=0 for population std
    std = np.std(fv, axis=0, ddof=1)
    z = (fv - mean) / std
    return z


def standardize_matrix(m):
    """Standardize a matrix using z-score."""
    if not isinstance(m, np.ndarray):
        m = np.array(m)
    mean = np.mean(m)
    # ddof=1 -> degrees of freedom = N-1, i.e. sample std
    # ddof = 'delta degrees of freedom'
    # set ddof=0 for population std
    std = np.std(m, ddof=1)
    z = (m - mean) / std
    return z


def normalize_matrix(m):
    """Normalize a matrix using min-max scaling."""
    m = np.asarray(m)
    norm = (m - m.min()) / (m.max() - m.min())
    return norm


def convert_cvrplib_to_vrp_instance(instance) -> VRPInstance:
    """Converts a `cvrplib.Instance.VRPTW` object to a
    `decomposition.VRPInstance` object.

    Parameters
    ----------
    instance: `cvrplib.Instance.VRPTW`
        A VRPTW problem instance returned by `cvrplib.read()`.

    Returns
    -------
    inst: `decomposition.VRPInstance`
        A `VRPInstance` object representing the VRP problem instance.

    """
    node_list = []
    for customer_id in range(len(instance.coordinates)):
        params = dict(
            x_coord = instance.coordinates[customer_id][0],
            y_coord = instance.coordinates[customer_id][1],
            demand = instance.demands[customer_id],
            distances = instance.distances[customer_id],
            start_time = instance.earliest[customer_id],
            end_time = instance.latest[customer_id],
            service_time = instance.service_times[customer_id],
        )
        node = Node(**params)
        node_list.append(node)

    inst = VRPInstance(node_list, instance.capacity,
        extra={'name': instance.name, 'num_vehicles': instance.n_vehicles})
    return inst


def get_time_window_overlap_or_gap(tw_1, tw_2):
    """Computes the amount of overlap or gap between 2 time windows.
        - Overlap: if overlap_or_gap > 0
        - Gap: if overlap_or_gap < 0

    For example:
        - let tw_1 = [0, 15], tw_2 = [5, 20], then there's an overlap of 10.
        - let tw_1 = [0, 10], tw_2 = [5, 9], then there's an overlap of 4.
        - let tw_1 = [0, 10], tw_2 = [15, 20], then there's a gap of 5.
    """
    tw_start_1, tw_end_1 = tw_1
    tw_start_2, tw_end_2 = tw_2
    overlap_or_gap = min(tw_end_1, tw_end_2) - max(tw_start_1, tw_start_2)
    return overlap_or_gap


def log_run_time(func: Callable):
    """Decorator for logging the run time of a function or method."""
    log = logger.getChild(func.__module__)

    def decorator(*args, **kwargs):
        start = time.time()
        return_val = func(*args, **kwargs)
        end = time.time()
        log.info(f'{func.__qualname__}() run time = {end-start} sec')
        return return_val

    return decorator


def df_to_excel(df: pd.DataFrame, file_name, sheet_name, overlay=True, index=False):
    file_name = file_name + '.xlsx'
    if_sheet_exists = 'overlay' if overlay else 'replace'
    start_row = 0
    header = True

    try:
        if overlay:
            # if sheet exists, append to the end of the same sheet
            wb = xl.load_workbook(file_name)
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # The maximum row index containing data (1-based)
                # `startrow` param of `pandas.DataFrame.to_excel()` is 0-based
                # So next row to append to is exactly max_row
                start_row = sheet.max_row
                # do not write out the column names again
                header = False
        # if sheet does not exist, create a new sheet in the existing excel file
        with pd.ExcelWriter(file_name, mode='a', if_sheet_exists=if_sheet_exists) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index, header=header, startrow=start_row)
    except FileNotFoundError:
        # if file does not yet exist, create it
        with pd.ExcelWriter(file_name, mode='x') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)


# alias for backward compatibility
write_to_excel = df_to_excel


def write_to_json(data, file_name, mode='a'):
    file_name = file_name + '.json'
    with open(file_name, mode) as f:
        json.dump(data, f) #, indent=4)
        f.write('\n')


def read_json(file_name):
    file_name = file_name + '.json'
    data = []
    # reading a file with multiple json objects must read 1 line/1 json object
    # at a time, bc json.load() can only handle a single json object
    with open(file_name) as f:
        for line in f:
            py_obj = json.loads(line)
            data.append(py_obj)

    return data


def read_json_gen(file_name):
    '''generator version of read_json'''
    file_name = file_name + '.json'
    # reading a file with multiple json objects must read 1 line/1 json object
    # at a time, bc json.load() can only handle a single json object
    with open(file_name) as f:
        for line in f:
            py_obj = json.loads(line)
            yield py_obj


def sleep(sec, logger_name=__name__):
    log = logger.getChild(logger_name)
    log.info(f'Sleeping for {sec} sec')
    time.sleep(sec)


def safe_divide(numerator, denominator):
    # prevent ZeroDivisionError by returning 0 if denominator is 0
    return numerator / denominator if denominator else 0


def make_dirs(dir_name):
    # make sure the parent directory exists
    dir_name = os.path.join(dir_name, 'placeholder_file_name')
    os.makedirs(os.path.dirname(dir_name), exist_ok=True)


def create_full_path_file_name(filename, *path_dir_names):
    path = os.path.join(*path_dir_names)
    make_dirs(path)
    fname = os.path.join(path, filename)
    return fname


class FV():
    """lru_cache requires function arguments to be hashable.
    Wrap a feature_vectors NDArray inside a user defined class
    to make it hashable for dist_matrix.
    """
    def __init__(self, data: np.ndarray, depot_data=None) -> None:
        self.data = data # customer data
        self.depot_data = depot_data


@lru_cache(maxsize=1)
def build_feature_vectors(inst: VRPInstance, standardize=False, include_depot=False):
    """Build feature vectors for clustering from VRP problem instance.
    A list of feature vectors representing the customer nodes
    to be clustered, excluding the depot.
    """
    fv_data = []
    nodes = inst.nodes
    for i in range(len(nodes)):
        row = []
        # x, y coords for customer i
        row.append(nodes[i].x_coord)
        row.append(nodes[i].y_coord)
        # earliest service start time for customer i
        row.append(nodes[i].start_time)
        # lastest service start time for customer i
        row.append(nodes[i].end_time)
        # service time for customer i
        row.append(nodes[i].service_time)

        fv_data.append(row)

    fv_data = np.asarray(fv_data)

    if standardize:
        fv_data = standardize_feature_vectors(fv_data)

    # By CVRPLIB convention, index 0 is always depot
    if include_depot:
        return FV(fv_data, fv_data[0])
    else:
        return FV(fv_data[1:], fv_data[0])


def get_hg_instance_names(n: int):
    '''Get all Homberger-Gehring instance names given size n in {2, 4, 6, 8, 10}'''
    types = ['C1', 'C2', 'R1', 'R2', 'RC1', 'RC2']
    for prefix in types:
        for i in range(1, 11):
            instance_name = f'{prefix}_{n}_{i}'
            yield instance_name


def read_instance(dir_name, instance_name, include_bk=False):
    file_name = os.path.join(CVRPLIB, dir_name, instance_name)
    inst, bk_sol = cvrplib.read(
        instance_path=f'{file_name}.txt',
        solution_path=f'{file_name}.sol'
    )
    converted_inst = convert_cvrplib_to_vrp_instance(inst)
    if include_bk:
        return inst, converted_inst, bk_sol
    else:
        return inst, converted_inst


def print_solution(solution, converted_inst, verbose=False):
    cost = solution.metrics[METRIC_COST]
    routes = solution.routes
    extra = solution.extra

    if extra is not None:
        route_starts = extra[EXTRA_ROUTE_STARTS]

    route_start = None
    total_wait_time = 0
    total_driving_time = 0
    for i, route in enumerate(routes):
        if extra is not None:
            route_start = route_starts[i]

        route_driving_time, route_wait_time = compute_route_time_and_wait_time(route, i, converted_inst, route_start, verbose=verbose)
        total_driving_time += route_driving_time
        total_wait_time += route_wait_time

    print(f'cost from solver: {cost}')
    print(f'computed total driving time: {total_driving_time}')
    print(f'computed total wait time: {total_wait_time}')
    print(f'extra: {extra}')
    print(f'num routes: {len(routes)}')
    print(f'routes: {routes}')

    return total_driving_time, total_wait_time


def compute_route_time_and_wait_time(route, route_num, inst: VRPInstance, route_start=None, verbose=False):
    # `route` doesn't include depot; `inst` does include depot = 0

    depot = 0
    first_stop = route[0]
    route_wait_time = 0
    route_driving_time = inst.nodes[depot].distances[first_stop]

    # don't count the wait time at the first stop
    # bc the vehicle could always be dispatched later from the depot
    # so as to not incur any wait time and it doesn't affect feasibility
    # NOTE: can't simply start at the earliest start time of the first node,
    # bc the earliest start time of the first node could be 0 and we can't
    # start at the first stop at time 0, bc we have to first travel from
    # deopt to the first stop
    if route_start is not None:
        depot_departure_time = route_start
    else:
        depot_departure_time = inst.nodes[depot].start_time + inst.nodes[depot].service_time
    # earliest possible arrival time at the first stop
    # = earliest possible time to leave the depot + driving time from depot to the first stop
    driving_time = inst.nodes[depot].distances[first_stop]
    arrival_time = depot_departure_time + driving_time
    # logical earliest start time at the first stop
    # = the later of arrival time and TW earliest start time
    tw_earliest_start = inst.nodes[first_stop].start_time
    tw_latest_start = inst.nodes[first_stop].end_time
    logical_earliest_start = max(arrival_time, tw_earliest_start)
    # departure time from the first node
    service_time = inst.nodes[first_stop].service_time
    departure_time = logical_earliest_start + service_time

    if verbose:
        print(f'--------------- START ROUTE {route_num + 1} ---------------')
        print(f'depot departure time = {depot_departure_time}')
        print(f'driving time from depot to {first_stop} = {driving_time}', end=', ')
        print(f'coords for depot = ({inst.nodes[depot].x_coord}, {inst.nodes[depot].y_coord})', end=', ')
        print(f'coords for stop {first_stop} = ({inst.nodes[first_stop].x_coord}, {inst.nodes[first_stop].y_coord})')

        print(f'stop = {first_stop}', end=', ')
        print(f'arrival time = {arrival_time}', end=', ')
        print(f'TW = [{tw_earliest_start}, {tw_latest_start}]', end=', ')
        print(f'logical earliest start = {logical_earliest_start}', end=', ')
        print(f'service time = {service_time}', end=', ')
        print(f'departure time = {departure_time}')
        print()

    prev_stop = first_stop
    for stop in route[1:]: # start counting wait time from the 2nd stop
        driving_time = inst.nodes[prev_stop].distances[stop]
        route_driving_time += driving_time
        tw_earliest_start = inst.nodes[stop].start_time
        tw_latest_start = inst.nodes[stop].end_time
        arrival_time = departure_time + driving_time
        # Wait if we arrive before earliest start
        wait_time = max(0, tw_earliest_start - arrival_time)
        route_wait_time += wait_time
        logical_earliest_start = arrival_time + wait_time
        service_time = inst.nodes[stop].service_time
        departure_time = logical_earliest_start + service_time

        if verbose:
            print(f'driving time from {prev_stop} to {stop} = {driving_time}', end=', ')
            print(f'coords for prev stop {prev_stop} = ({inst.nodes[prev_stop].x_coord}, {inst.nodes[prev_stop].y_coord})', end=', ')
            print(f'coords for stop {stop} = ({inst.nodes[stop].x_coord}, {inst.nodes[stop].y_coord})')

            print(f'stop = {stop}', end=', ')
            print(f'arrival time = {arrival_time}', end=', ')
            print(f'TW = [{tw_earliest_start}, {tw_latest_start}]', end=', ')
            print(f'wait time={wait_time}', end=', ')
            print(f'logical earliest start = {logical_earliest_start}', end=', ')
            print(f'service time = {service_time}', end=', ')
            print(f'departure time = {departure_time}')
            print()

        prev_stop = stop

    # back to the depot
    driving_time = inst.nodes[prev_stop].distances[depot]
    route_driving_time += driving_time
    arrival_time = departure_time + driving_time
    tw_earliest_start = inst.nodes[depot].start_time
    tw_latest_start = inst.nodes[depot].end_time

    if verbose:
        print(f'driving time from {prev_stop} to depot = {inst.nodes[prev_stop].distances[depot]}', end=', ')
        print(f'coords for prev stop {prev_stop} = ({inst.nodes[prev_stop].x_coord}, {inst.nodes[prev_stop].y_coord})', end=', ')
        print(f'coords for depot = ({inst.nodes[depot].x_coord}, {inst.nodes[depot].y_coord})')
        print(f'return time to depot = {arrival_time}', end=', ')
        print(f'depot TW = [{tw_earliest_start}, {tw_latest_start}]')
        print()

        print(f'route driving time = {route_driving_time}')
        print(f'route wait time = {route_wait_time}')
        print(f'--------------- END ROUTE {route_num + 1} ---------------')
        print()

    return route_driving_time, route_wait_time

