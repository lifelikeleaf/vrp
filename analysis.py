# Author: Xu Ye <kan.ye@tum.de>

import os
import numpy as np
from scipy.stats import norm, t
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, CellIsRule, FormulaRule
from openpyxl.utils.cell import get_column_letter
import vrp.decomp.helpers as helpers
from vrp.decomp.constants import *

num_subprobs_best_found = np.array([], dtype=int)

def get_best_found(df, name) -> pd.DataFrame:
    cost = df.loc[
        # 1. group by instance name
        # 2. get the index of the min cost within each group
        # 3. select only the rows that correspond to the min cost within each group
        df.groupby(by=KEY_INSTANCE_NAME)[KEY_COST].idxmin(), # rows
        [KEY_INSTANCE_NAME, KEY_NUM_SUBPROBS, KEY_COST] # cols
    ].sort_index().reset_index(drop=True)

    # a diff approach but same result
    # cost_wait = df.sort_values(by=[KEY_COST_WAIT]) \
    #     .groupby(by=KEY_INSTANCE_NAME).head(1).sort_index() \
    #     .loc[:, [KEY_INSTANCE_NAME, KEY_NUM_SUBPROBS, KEY_COST_WAIT]]

    # best_found = pd.merge(cost, cost_wait, on=[KEY_INSTANCE_NAME], suffixes=[f'_{KEY_COST}', f'_{KEY_COST_WAIT}'])

    global num_subprobs_best_found
    num_subprobs_best_found = np.append(num_subprobs_best_found, cost[f'{KEY_NUM_SUBPROBS}'].values)

    return cost


def get_avg(df):
    # this key function should be vectorized. It should expect a Series
    # and return a Series with the same shape as the input.
    # Thus it must use pandas.Series vectorized string functions for
    # extracting substrings and stripping underscores '_'.
    # Sort instance names by the ending digits, such that,
    # e.g. R1_10_3 would appear before R1_10_10.
    # Default python string sort would cause R1_10_10 to appear after R1_10_1
    # but before R1_10_2, like this: R1_10_1, R1_10_10, R1_10_2, R1_10_3, etc.
    # NOTE: this quick hack only works for HG instance names
    # i.e. C1_10_1, R1_2_10, RC1_10_9
    sort = lambda x: x.str.split('_', expand=True)[2].astype(int)
    # equivalent regex; select column 0 to convert type from DataFrame to Series
    # sort = lambda x: x.str.extract(r'\w\d_\d{1,2}_(\d{1,2})')[0].astype(int)

    # must use as_index=False, so the group by key (KEY_INSTANCE_NAME)
    # isn't used as index, otherwise it screws up column assignment
    # in comparision code, where comp['col_name'] = df[KEY_COST] - dfs['euc'][KEY_COST]
    # results in comp having NaN as the result of the column arithmitic
    avg = df.groupby(by=KEY_INSTANCE_NAME, as_index=False).mean() \
        .sort_values(by=KEY_INSTANCE_NAME, key=sort)[[KEY_INSTANCE_NAME, KEY_COST]] \
        .reset_index(drop=True)

    return avg


def get_variance(df):
    sort = lambda x: x.str.split('_', expand=True)[2].astype(int)
    var = df.groupby(by=KEY_INSTANCE_NAME, as_index=False).var() \
        .sort_values(by=KEY_INSTANCE_NAME, key=sort)[[KEY_INSTANCE_NAME, KEY_COST]] \
        .reset_index(drop=True)

    return var


def get_count(df):
    sort = lambda x: x.str.split('_', expand=True)[2].astype(int)
    count = df.groupby(by=KEY_INSTANCE_NAME, as_index=False).count() \
        .sort_values(by=KEY_INSTANCE_NAME, key=sort)[[KEY_INSTANCE_NAME, KEY_COST]] \
        .reset_index(drop=True)

    return count


def percent_diff(col1, col2, denom=None):
    if denom is None:
        # percent = (col1 - col2) / col2
        percent = np.divide(
            (col1 - col2),
            col2,
            out=np.zeros(col2.shape),
            where=(col2 != 0)
        )
    else:
        # percent = (col1 - col2) / denom
        percent = np.divide(
            (col1 - col2),
            denom,
            out=np.zeros(col2.shape),
            where=(denom != 0)
        )
    return percent


def dump_comparison_data(exp_names, dir_name, sub_dir, output_name, dump_best=False, dump_avg=False, dump_all=False, min_total=False):
    for exp_name in exp_names:
        input_file_name = os.path.join(dir_name, f'{dir_name}_{exp_name}.xlsx')
        dfs = dict(
            euc = pd.read_excel(input_file_name, sheet_name='euclidean'),
        )

        '''STEP 1/2 - Comparison'''
        '''MODIFY: sheet names and df column names'''

        '''for type C instances'''
        # dfs['qi_2012_0.99_0.01'] = pd.read_excel(input_file_name, sheet_name='qi_2012_0.99_0.01')

        # versions = ['v2_2_lambda_0.1']
        # for v in versions:
        #     dfs[f'OL_{v}'] = pd.read_excel(input_file_name, sheet_name=f'OL_{v}')

        '''for type R and RC instances'''
        dfs['qi_2012'] = pd.read_excel(input_file_name, sheet_name='qi_2012')

        versions = ['v2_2']
        for v in versions:
            dfs[f'Both_{v}'] = pd.read_excel(input_file_name, sheet_name=f'Both_{v}')

        '''END MODIFY'''

        basis = pd.read_excel(input_file_name, sheet_name='Basis')

        if dump_best:
            dfs_best = {name: get_best_found(df, name) for name, df in dfs.items()}
            comp_best = pd.DataFrame()
            comp_best[KEY_INSTANCE_NAME] = basis[KEY_INSTANCE_NAME]
            comp_best['euc vs nod'] = dfs_best['euc'][KEY_COST] - basis[f'{KEY_COST}_NO_decomp']
            comp_best['euc vs nod %'] = percent_diff(dfs_best['euc'][KEY_COST], basis[f'{KEY_COST}_NO_decomp'])
            # add an empty column
            comp_best[''] = ''
            output_name_best = f'best_{output_name}'
            dump_comp(dfs_best, comp_best, dir_name, sub_dir, output_name_best, exp_name, min_total, dump_best)

        if dump_avg:
            dfs_avg = {name: get_avg(df) for name, df in dfs.items()}
            comp_avg = pd.DataFrame()
            comp_avg[KEY_INSTANCE_NAME] = basis[KEY_INSTANCE_NAME]
            comp_avg['euc vs nod'] = dfs_avg['euc'][KEY_COST] - basis[f'{KEY_COST}_NO_decomp']
            comp_avg['euc vs nod %'] = percent_diff(dfs_avg['euc'][KEY_COST], basis[f'{KEY_COST}_NO_decomp'])
            # add an empty column
            comp_avg[''] = ''
            output_name_avg = f'avg_{output_name}'
            dump_comp(dfs_avg, comp_avg, dir_name, sub_dir, output_name_avg, exp_name, min_total)

        if dump_all:
            comp_all = pd.DataFrame()
            comp_all[KEY_INSTANCE_NAME] = dfs['euc'][KEY_INSTANCE_NAME]
            output_name_all = f'all_{output_name}'
            dump_comp(dfs, comp_all, dir_name, sub_dir, output_name_all, exp_name, min_total)


def dump_comp(dfs, comp, dir_name, sub_dir, output_name, sheet_name, min_total, dump_best=False):
    for name, df in dfs.items():
        if name != 'euc':
            comp[f'{name}_{KEY_COST}'] = df[KEY_COST] - dfs['euc'][KEY_COST]

    # cells containing the substring 'N/A' will be set to None below by formatting
    comp['N/A'] = ''

    for name, df in dfs.items():
        if name != 'euc':
            # percentage improvement compared to absolute euclidean cost
            comp[f'{name}_{KEY_COST}_%_euc'] = percent_diff(df[KEY_COST], dfs['euc'][KEY_COST])

    if not min_total: # calc "% improvement of improvement" only if OF = min driving time
        if dump_best or dump_avg:
            comp['N/A3'] = ''

            for name, df in dfs.items():
                if name != 'euc':
                    # percentage improvement compared to how much euclidean is able to improve no decomp
                    comp[f'{name}_{KEY_COST}_%_nod'] = percent_diff(df[KEY_COST], dfs['euc'][KEY_COST], abs(comp['euc vs nod']))

    dir_path = os.path.join(dir_name, sub_dir)
    helpers.make_dirs(dir_path)
    out = os.path.join(dir_path, output_name)
    helpers.df_to_excel(comp, file_name=out, sheet_name=sheet_name, overlay=False)


def conditional_formatting(dir_name, sub_dir, file_name):
    '''
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    c = ws['A4'] # a cell
    c.row # 4
    c.column # 1 == c.col_idx
    c.column_letter # 'A'
    '''
    file_to_load = os.path.join(dir_name, sub_dir, f'{file_name}.xlsx')
    wb = xl.load_workbook(file_to_load)

    for sheet in wb:
        # conditional color formatting
        green_fill = PatternFill(bgColor='00FF00', fill_type='solid')
        yellow_fill = PatternFill(bgColor='FFFF00', fill_type='solid')
        less_than_rule = CellIsRule(operator='lessThan', formula=['0'], fill=green_fill)
        # between_rule = CellIsRule(operator='between', formula=['-0.01', '0'], fill=yellow_fill) # use FormulaRule below
        '''Do not apply formula to blank cells, i.e. cell.value == None
        The formula uses a relative reference (A2), which will be expanded
        to the range over which the format is applied; this is a peculiar feature of Excel'''
        # apply if cell is not blank and cell value is in (-1%, 0]
        between_rule = FormulaRule(formula=['AND( NOT( ISBLANK(A2) ), A2 > -0.01, A2 <= 0 )'], fill=yellow_fill)
        start_row = sheet.min_row + 1 # skip the headers
        start_col = get_column_letter(sheet.min_column)
        end_row = sheet.max_row
        end_col = get_column_letter(sheet.max_column)
        # e.g. range_string = 'B2:F4'
        range_string = f'{start_col}{start_row}:{end_col}{end_row}'
        # order matters in conditional formatting: the first rule takes precedence
        # and is NOT overridden by later rules, so the more granular rule should be added first
        sheet.conditional_formatting.add(range_string, between_rule)
        sheet.conditional_formatting.add(range_string, less_than_rule)


        # make certain marked columns proper empty columns
        headers = sheet[1] # row 1
        for header_cell in headers:
            if header_cell.value is None or header_cell.value.strip() == '' or 'N/A' in header_cell.value:
                col = sheet[header_cell.column_letter]
                for i in range(len(col)):
                    cell = col[i]
                    cell.value = None


        # add additional rows for counting comparison results per column
        cur_max_row = sheet.max_row
        added_row1 = cur_max_row + 2
        added_row2 = cur_max_row + 3
        added_row3 = cur_max_row + 4
        cols_gen = sheet.columns
        for col in cols_gen:
            cell = col[0]
            if cell.value is not None:
                col_letter = cell.column_letter
                cell_loc_1 = f'{col_letter}{added_row1}'
                cell_loc_2 = f'{col_letter}{added_row2}'
                cell_loc_3 = f'{col_letter}{added_row3}'
                formula_range = f'{col_letter}{start_row}:{col_letter}{cur_max_row}'
                if col_letter == 'A': # first column
                    sheet[cell_loc_1] = 'x<0'
                    # sheet[cell_loc_2] = '0<=x<=0.01'
                    # sheet[cell_loc_3] = 'x<=0.01'
                    sheet[cell_loc_2] = 'sum <0'
                    sheet[cell_loc_3] = 'sum all'
                else:
                    sheet[cell_loc_1] = f'=COUNTIF({formula_range}, "<0")'
                    # sheet[cell_loc_2] = f'=COUNTIFS({formula_range}, ">=0", {formula_range}, "<=0.01")'
                    # sheet[cell_loc_3] = f'=COUNTIF({formula_range}, "<=0.01")'
                    sheet[cell_loc_2] = f'=SUMIF({formula_range}, "<0")'
                    sheet[cell_loc_3] = f'=SUM({formula_range})'


        # format % columns
        headers = sheet[1] # row 1; 1-based indexing of excel
        for header_cell in headers:
            if header_cell.value is not None and '%' in header_cell.value:
                percent_col = sheet[header_cell.column_letter]
                for i in range(1, len(percent_col)): # skip the header row; 0-based indexing of column tuple
                    if i != added_row1 - 1: # minus 1 due to 0-based indexing of column tuple and 1-based indexing of excel
                        cell = percent_col[i]
                        cell.number_format = '0.00%'


        # conditional format the above added rows one block at a time
        pink_fill = PatternFill(bgColor='FFCCCC', fill_type='solid')
        orange_fill = PatternFill(bgColor='FFCC33', fill_type='solid')
        rank = 2
        top_rule = Rule(type='top10', rank=rank, dxf=DifferentialStyle(fill=pink_fill))
        bottom_rule = Rule(type='top10', bottom=True, rank=rank, dxf=DifferentialStyle(fill=orange_fill))

        cols_gen = sheet.columns
        start_col = 'B' # skip column A, which contains instance names
        for col in cols_gen:
            cell = col[0]
            if cell.value is None:
                # the column before the empty column is the end column for this block
                end_col = get_column_letter(cell.column - 1)
                sheet.conditional_formatting.add(f'{start_col}{added_row1}:{end_col}{added_row1}', top_rule)
                sheet.conditional_formatting.add(f'{start_col}{added_row2}:{end_col}{added_row2}', bottom_rule)
                sheet.conditional_formatting.add(f'{start_col}{added_row3}:{end_col}{added_row3}', bottom_rule)
                # new start column is the column after the empty column
                start_col = get_column_letter(cell.column + 1)

        # handle the final block after cols_gen has been exhausted
        end_col = get_column_letter(cell.column)
        sheet.conditional_formatting.add(f'{start_col}{added_row1}:{end_col}{added_row1}', top_rule)
        sheet.conditional_formatting.add(f'{start_col}{added_row2}:{end_col}{added_row2}', bottom_rule)
        sheet.conditional_formatting.add(f'{start_col}{added_row3}:{end_col}{added_row3}', bottom_rule)


    dir_path = os.path.join(dir_name, sub_dir)
    helpers.make_dirs(dir_path)
    out = os.path.join(dir_path, f'{file_name}_formatted.xlsx')
    wb.save(out)


def calc_confidence_interval(df, alpha=0.05):
    n = get_count(df)[KEY_COST][0]
    probability = 1 - alpha / 2
    # z = norm.ppf(probability) # NORM.S.INV(prob): inverse of cdf
    z = t.ppf(probability, n - 1) # use t distribution; T.INV(prob, dof)
    ci = pd.DataFrame()
    avg = get_avg(df)
    ci[KEY_INSTANCE_NAME] = avg[KEY_INSTANCE_NAME]
    ci['mean'] = avg[KEY_COST]
    ci['variance'] = get_variance(df)[KEY_COST]
    ci['count'] = get_count(df)[KEY_COST]
    ci['variance_of_mean'] = ci['variance'] / ci['count']
    ci['half_width'] = z * np.sqrt(ci['variance_of_mean'])
    ci['low'] = ci['mean'] - ci['half_width']
    ci['high'] = ci['mean'] + ci['half_width']
    return ci


def calc_confidence_intervals(exp_names, dir_name, sub_dir, alpha=0.05):
    for exp_name in exp_names:
        input_file_name = os.path.join(dir_name, f'{dir_name}_{exp_name}.xlsx')
        dfs = dict(
            euc = pd.read_excel(input_file_name, sheet_name='euclidean'),
        )

        '''STEP 1/2 - Confidence Interval'''
        '''MODIFY: sheet names and df column names'''

        '''for type C instances only'''
        # dfs['qi_2012_0.99_0.01'] = pd.read_excel(input_file_name, sheet_name='qi_2012_0.99_0.01')

        # versions = ['v2_2_lambda_0.1']
        # for v in versions:
        #     dfs[f'OL_{v}'] = pd.read_excel(input_file_name, sheet_name=f'OL_{v}')

        '''for all instances, but particularly for type R and RC instances'''
        dfs['qi_2012'] = pd.read_excel(input_file_name, sheet_name='qi_2012')

        versions = ['v2_2']
        for v in versions:
            dfs[f'Both_{v}'] = pd.read_excel(input_file_name, sheet_name=f'Both_{v}')

        '''END MODIFY'''

        for name, df in dfs.items():
            # ci = calc_confidence_interval(df, alpha)
            # out = helpers.create_full_path_file_name(name, dir_name, sub_dir, 'confidence_intervals')
            # helpers.df_to_excel(ci, file_name=out, sheet_name=exp_name, overlay=False)
            # continue

            # compare 2 systems (using Paired-t Confidence Interval)
            ref = 'euc' # reference column for comparison
            # ref = 'qi_2012'
            if name != ref and name != 'euc':
                filename = f'diff_{ref}_{name}'
                diff = pd.DataFrame()
                diff[KEY_INSTANCE_NAME] = df[KEY_INSTANCE_NAME]
                diff[KEY_COST] = df[KEY_COST] - dfs[ref][KEY_COST]

                raw = calc_confidence_interval(diff, alpha)

                ci = raw.loc[:, [KEY_INSTANCE_NAME, 'mean', 'half_width']]
                # format a number with a 1000 separator and 2 decimal places,
                # pad it with '[' on the left and ', ' on the right
                # so 567.89123, for example, becomes '[567.89, '
                lo = raw['low'].map('[{:,.2f}, '.format)
                # pad it with ']' on the right
                # so 1234.5678, for example, becomes '1,234.57]'
                hi = raw['high'].map('{:,.2f}]'.format)
                ci.loc[:, 'interval'] = lo + hi # final form looks like '[567.89, 1,234.57]'
                ci.loc[:, ['low', 'high']] = raw.loc[:, ['low', 'high']]

                out = helpers.create_full_path_file_name(filename, dir_name, sub_dir, 'confidence_intervals')
                helpers.df_to_excel(ci, file_name=out, sheet_name=exp_name, overlay=False)


if __name__ == '__main__':
    '''STEP 2/2'''
    '''MODIFY: experiment names and dir name'''

    dir_name = 'E28'
    min_total = True
    print_num_subprobs = False # this is only meaningful if dump_best = True
    dump_best = False
    dump_avg = True
    dump_all = False

    exp = 'k_medoids'
    exp_names = [
        f'{exp}_C1',
        f'{exp}_C2',
        f'{exp}_R1',
        f'{exp}_R2',
        f'{exp}_RC1',
        f'{exp}_RC2',
    ]

    '''END MODIFY'''

    sub_dir = file_name = f'{dir_name}_comparison'

    calc_confidence_intervals(exp_names, dir_name, sub_dir)
    exit()

    dump_comparison_data(exp_names, dir_name, sub_dir, file_name, dump_best=dump_best, dump_avg=dump_avg, dump_all=dump_all, min_total=min_total)
    if print_num_subprobs:
        print(num_subprobs_best_found.tolist())
        for i in range(num_subprobs_best_found.min(), num_subprobs_best_found.max() + 1, 1):
            print(f'num_clusters={i}: {len(num_subprobs_best_found[num_subprobs_best_found == i])}')
        fig, ax = plt.subplots()
        ax.hist(num_subprobs_best_found, bins=len(num_subprobs_best_found), linewidth=0.5, edgecolor="white", align='mid', rwidth=0.8)
        plt.show()

    if dump_best:
        conditional_formatting(dir_name, sub_dir, f'best_{file_name}')
    if dump_avg:
        conditional_formatting(dir_name, sub_dir, f'avg_{file_name}')
    if dump_all:
        conditional_formatting(dir_name, sub_dir, f'all_{file_name}')

